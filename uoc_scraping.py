#!/usr/bin/env python3.5
from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles.borders import Border, Side
import warnings

''' 
data breakdown (relevant for output of scrape_document/input for export data):
data[0] = unit code e.g. UEENEEK142A
data[1] = unit name e.g. Apply environmentally and sustainable procedures in the energy sector
data[2] = assessment tool version
data[3] = range statement
data[4] = ciritcal aspect of evidence
data[5] = elements and performance criteria
data[6] = required skill and knowledge
'''

def scrape_document(uoc):
    active = 0 #used to 'turn on/off' features
    page = requests.get("http://training.gov.au/Training/Details/" + uoc)

    soup = BeautifulSoup(page.content,'html.parser')
    document_title = soup.title.get_text()

    unit_code = uoc
    unit_name = document_title.split('-')[-1].lstrip()
    assess_tool_ver = uoc + "_Assessment_Mapping_Toolv1.1"
    range_statement = ""
    asp_of_evidence = ""
    elem_perf_criteria = ""
    req_skill_knowledge = ""

    for e in soup.find_all("h2"):
        if e.string == "Range Statement":
            range_table = e.find_next_siblings()[0]
            range_statement = extract_table_data(range_table)

        if e.string == "Evidence Guide":
            evidence_guide_table = e.find_next_siblings()[0]          
            asp_of_evidence = extract_table_data(evidence_guide_table)
            start = asp_of_evidence.index("Critical aspects of evidence required to demonstrate competency in this unit")
            end = asp_of_evidence.index("Context of and specific resources for assessment")
            asp_of_evidence = asp_of_evidence[start:end]

        if e.string == "Elements and Performance Criteria" and active:
            perf_criteria_table = e.find_next_siblings()[0]
            data = (perf_criteria_table.find_all("tr"))
            for i in range(len(data)):
                elem_perf_criteria += data[i].get_text()

        if e.string == "Required Skills and Knowledge" and active:
            k_and_s_table = e.find_next_siblings()[0]
            data = (k_and_s_table.find_all("tr"))
            for i in range(len(data)):
                req_skill_knowledge += data[i].get_text()
    return [unit_code, unit_name, assess_tool_ver, range_statement,
            asp_of_evidence, elem_perf_criteria, req_skill_knowledge]


def extract_table_data(table):
    s = ''
    for row in table.find_all("td"):
        s += row.get_text()
    return s

def export_data(data, input_location, output_location):
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    book = openpyxl.load_workbook(input_location)
    curr_sheet = book.get_sheet_by_name("Sheet1")
    curr_sheet.cell(row = 3, column = 2).value = data[0]
    curr_sheet.cell(row = 4, column = 2).value = data[1]
    curr_sheet.cell(row = 5, column = 2).value = "UEE11"
    curr_sheet.cell(row = 7, column = 2).value = data[2]
    curr_sheet.cell(row = 8, column = 2).value = data[3]
    curr_sheet.cell(row = 9, column = 2).value = data[4]
    curr_sheet.cell(row = 13, column = 1).value = data[5]
    curr_sheet.cell(row = 15, column = 1).value = data[0] + ' - ' + data[1]
    curr_sheet.cell(row = 45, column = 1).value = data[6]

    for row in range(2,10):
        thicken_border(2,8, row, curr_sheet, thin_border)
    thicken_border(1,9, 15, curr_sheet, thin_border)

    output_file = output_location + data[2] + "MODIFIED.xlsx"
    book.save(output_file)

def thicken_border(start,end, row, curr_sheet, style):
    for i in range(start,end):
        curr_sheet.cell(row = row, column = i).border = style


document_title = ""

unit = input("Enter Unit of Competency Code e.g. UEENEEK142A: ")

warnings.filterwarnings("ignore")
data = scrape_document(unit)

print("Enter file path to location where template is stored")
print("Example: C:/Users/harri/Documents/UEXXXXXXX_Assessment_Mapping_Tool v1.1 MASTER PH.xlsx")
template_location = input("Location: ")

print("Enter file path to location where output is stored")
print("Example: C:/Users/harri/Documents/")
output_location = input("Location: ")
export_data(data, template_location, output_location)

print("Process Complete")
